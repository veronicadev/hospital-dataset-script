
import os 
import re
import sys
import unidecode
from openpyxl import load_workbook
import requests

ACCESS_KEY = os.environ.get('ACCESS_KEY')

class HospitalsDatasetConverter():
    def get_coordinates(self, address):
        coordinates = ""
        response = requests.get("http://api.positionstack.com/v1/forward?access_key="+ACCESS_KEY+"&query="+address)
        response_json = response.json()
        if response_json["data"]:
            if response_json["data"][0]:
                coordinates = str(response_json["data"][0]["latitude"]) + ","+ str(response_json["data"][0]["longitude"])
        return coordinates

    def elaborate(self, hospitals_path, name, output_name):
        #load excel file
        hospitals_xlsx = load_workbook(filename=hospitals_path+name)
        for sheet in hospitals_xlsx:
            # se non è presente la colonna entity_id e coordinates, la creo in colonna D e E
            # se D1 e E1 sono entity_id e coordinates procedo
            print("Elaborating: "+sheet.title)

            dCol = sheet.cell(row=1, column=4)
            eCol = sheet.cell(row=1, column=5)
            if dCol.value != "coordinates":
                sheet.insert_cols(5);
                dCol.value = "coordinates"

            if eCol.value != "entity_id":
                sheet.insert_cols(6);
                eCol.value = "entity_id"

            rows = tuple(sheet.rows)
            i = 1;
            while i < len(rows) and rows[i][1].value:
                entity_id = re.sub(r'[^\w]', ' ', str(rows[i][2].value))
                entity_id = re.sub(' +', ' ', str(entity_id)).strip()
                entity_id = re.sub(r'[^\w]', '_', str(entity_id)).upper()
                rows[i][4].value = unidecode.unidecode(entity_id)
                rows[i][3].value = self.get_coordinates(rows[i][10].value+", "+rows[i][9].value)
                #print(rows[i][4].value)
                #print("\n")
                i = i+1

            
        #save the file
        hospitals_xlsx.save(filename=hospitals_path+output_name)



def main():
    if len(sys.argv)>3:
        print("++++++++++++++++++++++++++++")
        print("HospitalsDatasetConverter: START" )
        xlsx = HospitalsDatasetConverter()
        xlsx.elaborate(sys.argv[1], sys.argv[2], sys.argv[3])
        print("HospitalsDatasetConverter: END" )
        print("++++++++++++++++++++++++++++")
    else:
        print("Please provide valid parameters!")

if __name__ == "__main__":
    main()